# -*- coding: utf-8 -*-
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

src = Path(r"c:\Users\Saleem\OneDrive\Desktop\سجل المشتركين 2026 محدث\كشف راوترات الشبكة.xlsx")
tpl = Path(r"e:\files2\راوترات_الشبكة_قالب_فارغ.xlsx")

print("Source exists:", src.exists(), file=sys.stderr)
print("Template exists:", tpl.exists(), file=sys.stderr)

if src.exists():
    wb = openpyxl.load_workbook(src, data_only=True)
    print("Source sheets:", wb.sheetnames)
    ws = wb.active
    print("Active:", ws.title, "rows:", ws.max_row, "cols:", ws.max_column)
    port_rows = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 60)):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and "Port" in v:
                port_rows.append((r, c, v))
    print("Port markers:", len(port_rows))
    for pr in port_rows:
        print(f"  R{pr[0]}: {pr[2]}")
    if port_rows:
        r = port_rows[0][0]
        print(f"--- Sample data after first port (row {r+1}) ---")
        for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
            parts = []
            for cc in range(1, min(35, ws.max_column + 1)):
                v = ws.cell(rr, cc).value
                if v is not None and str(v).strip():
                    letter = openpyxl.utils.get_column_letter(cc)
                    parts.append(f"{letter}={str(v)[:30]}")
            print(f"  R{rr}: {' | '.join(parts)}")

if tpl.exists():
    wb2 = openpyxl.load_workbook(tpl, data_only=False)
    print("Template sheets:", wb2.sheetnames)
    for sn in wb2.sheetnames[:3]:
        ws2 = wb2[sn]
        print(f"--- {sn} ---")
        for r in range(1, 5):
            vals = [ws2.cell(r, c).value for c in range(1, 10)]
            print(f"  R{r}: {vals}")
