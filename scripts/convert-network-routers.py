#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحويل «كشف راوترات الشبكة.xlsx» إلى قالب MASH ISP.

القواعد الصحيحة:
  - كل قسم Port يُكتب في تبويبه الفعلي (Port 2 → 2️⃣ Port 2، وليس Port 1).
  - صف واحد = راوتر رئيسي + موسّع (أعمدة J/K) إن وُجد.
  - أجهزة Bypass → ملف Bypassed منفصل.

تشغيل:
  py -m pip install openpyxl
  py scripts/convert-network-routers.py
"""

from __future__ import annotations

import re
import shutil
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

SOURCE_PATH = Path(
    r"c:\Users\Saleem\OneDrive\Desktop\سجل المشتركين 2026 محدث\كشف راوترات الشبكة.xlsx"
)
TEMPLATE_PATH = Path(r"e:\files2\راوترات_الشبكة_قالب_فارغ.xlsx")
BYPASSED_TEMPLATE_PATH = Path(
    r"E:\Project\mash-isp\public\templates\network-bypassed-template.xlsx"
)
OUTPUT_PATH = TEMPLATE_PATH.parent / "راوترات_الشبكة_مملوء.xlsx"
BYPASSED_OUTPUT_PATH = TEMPLATE_PATH.parent / "نموذج_Bypassed_مملوء.xlsx"

ROUTER_DATA_START_ROW = 4
ROUTER_MAX_ROWS = 254
BYPASSED_DATA_START_ROW = 2

TEMPLATE_PORT_TABS: dict[int, str] = {
    1: "1️⃣ Port 1",
    2: "2️⃣ Port 2",
    3: "3️⃣ Port 3",
    4: "4️⃣ Port 4",
    5: "5️⃣ Port 5",
    6: "6️⃣ Port 6",
    7: "7️⃣ Port 7",
    8: "8️⃣ Port 8",
    9: "9️⃣ Port 9",
}

SRC_MAIN = {
    "ip": 3,
    "code": 4,
    "mac": 5,
    "location": 6,
    "ssid": 7,
    "device_type": 8,
    "phone": 10,
    "phone_alt": 9,
}

SRC_EXTENDER = {
    "name": 14,
    "mac": 15,
    "port": 16,
    "ssid": 17,
    "phone": 18,
}

SRC_BYPASS1 = {
    "name": 22,
    "mac": 23,
    "port": 24,
    "device_type": 25,
    "extra": 26,
}

SRC_BYPASS2 = {
    "ssid": 29,
    "ip": 30,
    "mac": 31,
    "device_type": 32,
    "notes": 33,
}

PORT_SECTION_RE = re.compile(r"Port\s*(\d+)\s*[-–—]?\s*(Router|Switch)", re.I)


def cell_value(ws: Worksheet, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def as_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def looks_like_ip(s: str) -> bool:
    return bool(re.match(r"^\d{1,3}(\.\d{1,3}){3}$", s))


def looks_like_mac(s: str) -> bool:
    return bool(re.match(r"^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$", s))


def looks_like_phone(s: str) -> bool:
    digits = re.sub(r"\D", "", s)
    return len(digits) >= 7 and not looks_like_ip(s) and not looks_like_mac(s)


def build_notes(kind: str, port: str, extras: list[str]) -> str:
    parts: list[str] = [kind]
    if port:
        parts.append(f"PORT={port}")
    parts.extend(x for x in extras if x)
    return " | ".join(parts)


def extract_main(ws: Worksheet, row: int) -> dict[str, str] | None:
    ip = as_str(cell_value(ws, row, SRC_MAIN["ip"]))
    code = as_str(cell_value(ws, row, SRC_MAIN["code"]))
    mac = as_str(cell_value(ws, row, SRC_MAIN["mac"]))
    location = as_str(cell_value(ws, row, SRC_MAIN["location"]))
    ssid = as_str(cell_value(ws, row, SRC_MAIN["ssid"]))
    device_type = as_str(cell_value(ws, row, SRC_MAIN["device_type"]))
    phone = as_str(cell_value(ws, row, SRC_MAIN["phone"])) or as_str(
        cell_value(ws, row, SRC_MAIN["phone_alt"])
    )

    if not any([ip, code, mac, location, ssid]):
        return None

    return {
        "ip": ip,
        "code": code,
        "mac": mac,
        "location": location,
        "ssid": ssid,
        "device_type": device_type,
        "phone": phone,
        "notes": "",
    }


def extract_router_row(ws: Worksheet, row: int) -> dict[str, str] | None:
    """صف واحد في القالب = راوتر رئيسي + موسّع في نفس الصف."""
    main = extract_main(ws, row)
    if not main:
        return None

    ext_name = as_str(cell_value(ws, row, SRC_EXTENDER["name"]))
    ext_mac = as_str(cell_value(ws, row, SRC_EXTENDER["mac"]))
    ext_port = as_str(cell_value(ws, row, SRC_EXTENDER["port"]))
    ext_ssid = as_str(cell_value(ws, row, SRC_EXTENDER["ssid"]))

    notes: list[str] = []
    if ext_port:
        notes.append(f"موسّع PORT={ext_port}")
    if ext_ssid and ext_ssid != ext_name:
        notes.append(f"SSID موسّع: {ext_ssid}")

    return {
        **main,
        "notes": " | ".join(notes),
        "ext_name": ext_name,
        "ext_mac": ext_mac,
    }


def extract_bypass1(ws: Worksheet, row: int, main_ssid: str) -> dict[str, str] | None:
    name = as_str(cell_value(ws, row, SRC_BYPASS1["name"]))
    mac = as_str(cell_value(ws, row, SRC_BYPASS1["mac"]))
    port = as_str(cell_value(ws, row, SRC_BYPASS1["port"]))
    device_type = as_str(cell_value(ws, row, SRC_BYPASS1["device_type"])) or "Bypassed"
    extra = as_str(cell_value(ws, row, SRC_BYPASS1["extra"]))

    ip = extra if looks_like_ip(extra) else ""
    phone = extra if looks_like_phone(extra) else ""
    extra_note = extra if extra and not ip and not phone else ""

    if not any([name, mac, ip]):
        return None

    extras = [f"راوتر رئيسي: {main_ssid}"] if main_ssid else []
    if extra_note:
        extras.append(extra_note)

    return {
        "ip": ip,
        "code": "",
        "mac": mac,
        "location": name,
        "ssid": name,
        "device_type": device_type,
        "phone": phone,
        "notes": build_notes("Bypassed", port, extras),
    }


def extract_bypass2(ws: Worksheet, row: int, main_ssid: str) -> dict[str, str] | None:
    ssid = as_str(cell_value(ws, row, SRC_BYPASS2["ssid"]))
    ip = as_str(cell_value(ws, row, SRC_BYPASS2["ip"]))
    mac = as_str(cell_value(ws, row, SRC_BYPASS2["mac"]))
    device_type = as_str(cell_value(ws, row, SRC_BYPASS2["device_type"])) or "Bypassed"
    notes_raw = as_str(cell_value(ws, row, SRC_BYPASS2["notes"]))

    if not any([ssid, ip, mac]):
        return None

    extras = [f"راوتر رئيسي: {main_ssid}"] if main_ssid else []
    if notes_raw:
        extras.append(notes_raw)

    return {
        "ip": ip,
        "code": "",
        "mac": mac,
        "location": ssid,
        "ssid": ssid,
        "device_type": device_type,
        "phone": "",
        "notes": build_notes("Bypassed", "", extras),
    }


def parse_port_from_title(title: str) -> int | None:
    m = PORT_SECTION_RE.search(title)
    if not m:
        return None
    n = int(m.group(1))
    return n if 1 <= n <= 9 else None


def find_port_sections(ws: Worksheet) -> list[tuple[int, str, int]]:
    seen: set[int] = set()
    sections: list[tuple[int, str, int]] = []

    for row in range(1, ws.max_row + 1):
        if row in seen:
            continue
        title = as_str(cell_value(ws, row, 1))
        port_num = parse_port_from_title(title) if title else None
        if port_num is not None:
            sections.append((row, title, port_num))
            seen.add(row)

    sections.sort(key=lambda x: x[0])
    return sections


def extract_section(
    ws: Worksheet, start_row: int, end_row: int
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    routers: list[dict[str, str]] = []
    bypasses: list[dict[str, str]] = []

    for row in range(start_row + 1, end_row):
        first_col = as_str(cell_value(ws, row, 1))
        if first_col and parse_port_from_title(first_col):
            continue

        router = extract_router_row(ws, row)
        if not router:
            continue

        routers.append(router)
        main_ssid = router["ssid"]

        for extractor in (extract_bypass1, extract_bypass2):
            bp = extractor(ws, row, main_ssid)
            if bp:
                bypasses.append(bp)

    return routers, bypasses


def clear_router_data(ws: Worksheet) -> None:
    for row in range(ROUTER_DATA_START_ROW, ROUTER_DATA_START_ROW + ROUTER_MAX_ROWS):
        for col in range(2, 12):  # B..K
            ws.cell(row=row, column=col).value = None


def write_routers_to_sheet(ws: Worksheet, records: list[dict[str, str]]) -> int:
    clear_router_data(ws)
    overflow = 0

    for i, rec in enumerate(records, start=1):
        if i > ROUTER_MAX_ROWS:
            overflow += 1
            continue
        row = ROUTER_DATA_START_ROW + i - 1
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = rec["ip"] or None
        ws.cell(row=row, column=3).value = rec["code"] or None
        ws.cell(row=row, column=4).value = rec["mac"] or None
        ws.cell(row=row, column=5).value = rec["location"] or None
        ws.cell(row=row, column=6).value = rec["ssid"] or None
        ws.cell(row=row, column=7).value = rec["device_type"] or None
        ws.cell(row=row, column=8).value = rec["phone"] or None
        ws.cell(row=row, column=9).value = rec["notes"] or None
        ws.cell(row=row, column=10).value = rec.get("ext_name") or None
        ws.cell(row=row, column=11).value = rec.get("ext_mac") or None

    written = min(len(records), ROUTER_MAX_ROWS)
    for row in range(ROUTER_DATA_START_ROW + written, ROUTER_DATA_START_ROW + ROUTER_MAX_ROWS):
        ws.cell(row=row, column=1).value = row - ROUTER_DATA_START_ROW + 1

    return overflow


def write_bypassed_sheet(ws: Worksheet, records: list[dict[str, str]]) -> None:
    max_row = ws.max_row
    for row in range(BYPASSED_DATA_START_ROW, max_row + 1):
        for col in range(1, 10):
            ws.cell(row=row, column=col).value = None

    for i, rec in enumerate(records, start=1):
        row = BYPASSED_DATA_START_ROW + i - 1
        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = rec["ip"] or None
        ws.cell(row=row, column=3).value = rec["code"] or None
        ws.cell(row=row, column=4).value = rec["mac"] or None
        ws.cell(row=row, column=5).value = rec["location"] or None
        ws.cell(row=row, column=6).value = rec["ssid"] or None
        ws.cell(row=row, column=7).value = rec["device_type"] or None
        ws.cell(row=row, column=8).value = rec["phone"] or None
        ws.cell(row=row, column=9).value = rec["notes"] or None


def main() -> int:
    if not SOURCE_PATH.exists():
        print(f"ERROR: ملف المصدر غير موجود: {SOURCE_PATH}", file=sys.stderr)
        return 1
    if not TEMPLATE_PATH.exists():
        print(f"ERROR: ملف القالب غير موجود: {TEMPLATE_PATH}", file=sys.stderr)
        return 1

    print("قراءة المصدر...")
    src_ws = openpyxl.load_workbook(SOURCE_PATH, data_only=True).worksheets[0]
    sections = find_port_sections(src_ws)
    print(f"   أقسام Port: {len(sections)}")
    for row, title, port_num in sections:
        print(f"   - صف {row}: {title} -> {TEMPLATE_PORT_TABS[port_num]}")

    routers_by_port: dict[int, list[dict[str, str]]] = {}
    all_bypasses: list[dict[str, str]] = []

    for idx, (start_row, title, port_num) in enumerate(sections):
        end_row = sections[idx + 1][0] if idx + 1 < len(sections) else src_ws.max_row + 1
        routers, bypasses = extract_section(src_ws, start_row, end_row)
        routers_by_port[port_num] = routers
        all_bypasses.extend(bypasses)
        print(f"   Port {port_num}: {len(routers)} راوتر، {len(bypasses)} bypass")

    print(f"\nنسخ القالب -> {OUTPUT_PATH}")
    output_path = OUTPUT_PATH
    try:
        shutil.copy2(TEMPLATE_PATH, output_path)
    except PermissionError:
        output_path = OUTPUT_PATH.with_stem(OUTPUT_PATH.stem + "_v2")
        print(f"WARN: الملف الأصلي مفتوح — الحفظ في: {output_path}")
        shutil.copy2(TEMPLATE_PATH, output_path)
    out_wb = openpyxl.load_workbook(output_path)

    total_routers = 0
    for port_num, records in sorted(routers_by_port.items()):
        sheet_name = TEMPLATE_PORT_TABS[port_num]
        if sheet_name not in out_wb.sheetnames:
            print(f"ERROR: تبويب غير موجود: {sheet_name}", file=sys.stderr)
            return 1
        overflow = write_routers_to_sheet(out_wb[sheet_name], records)
        total_routers += len(records)
        msg = f"OK {sheet_name}: {len(records)} راوتر"
        if overflow:
            msg += f" (WARN: {overflow} لم يُكتب — تجاوز 254)"
        print(msg)

    out_wb.save(output_path)

    bypassed_output_path = BYPASSED_OUTPUT_PATH
    if all_bypasses and BYPASSED_TEMPLATE_PATH.exists():
        print(f"\nكتابة Bypassed -> {bypassed_output_path}")
        try:
            shutil.copy2(BYPASSED_TEMPLATE_PATH, bypassed_output_path)
        except PermissionError:
            bypassed_output_path = BYPASSED_OUTPUT_PATH.with_stem(BYPASSED_OUTPUT_PATH.stem + "_v2")
            print(f"WARN: الملف الأصلي مفتوح — الحفظ في: {bypassed_output_path}")
            shutil.copy2(BYPASSED_TEMPLATE_PATH, bypassed_output_path)
        bp_wb = openpyxl.load_workbook(bypassed_output_path)
        write_bypassed_sheet(bp_wb.active, all_bypasses)
        bp_wb.save(bypassed_output_path)
        print(f"OK Bypassed: {len(all_bypasses)} جهاز")
    elif all_bypasses:
        print(f"WARN: {len(all_bypasses)} bypass — قالب Bypassed غير موجود")

    print(f"\nتم الحفظ:")
    print(f"  - {output_path} ({total_routers} راوتر)")
    if all_bypasses and BYPASSED_TEMPLATE_PATH.exists():
        print(f"  - {bypassed_output_path} ({len(all_bypasses)} bypass)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
